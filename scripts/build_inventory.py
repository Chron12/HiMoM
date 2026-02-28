#!/usr/bin/env python3
"""
Build the master inventory spreadsheet for PreRollTracker and ApexAPI.
Generates /Users/chrisgillis/PycharmProjects/HiMoM/docs/inventory.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADER_FILL = PatternFill(start_color="1A5676", end_color="1A5676", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
BODY_FONT = Font(name="Calibri", size=10)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="DDDDDD"),
)

# Sheet tab colors
TAB_COLORS = {
    "Pages & Screens": "1A5676",
    "API Endpoints": "2E86AB",
    "Database Tables": "A23B72",
    "Background Jobs": "F18F01",
    "Integrations": "C73E1D",
    "Config Options": "3B1F2B",
}

MAX_COL_WIDTH = 55
MIN_COL_WIDTH = 12


def style_sheet(ws, num_cols):
    """Apply professional formatting to a worksheet."""
    # Freeze first row
    ws.freeze_panes = "A2"

    # Style header row
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Style data rows with alternating colors
    for row_idx in range(2, ws.max_row + 1):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGNMENT
            cell.border = THIN_BORDER

    # Auto-width columns
    for col_idx in range(1, num_cols + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                # Take the longest line in the cell for width calculation
                lines = str(val).split("\n")
                longest_line = max(len(line) for line in lines)
                max_length = max(max_length, longest_line)
        width = min(max(max_length + 3, MIN_COL_WIDTH), MAX_COL_WIDTH)
        ws.column_dimensions[col_letter].width = width

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{ws.max_row}"


def add_rows(ws, headers, data):
    """Add header row and data rows to a worksheet."""
    ws.append(headers)
    for row in data:
        ws.append(row)


def build_pages_sheet(ws):
    headers = ["App", "URL / Location", "Page Name", "Purpose", "Primary Audience", "Auth Required"]
    data = [
        # PreRollTracker pages
        ["PreRollTracker", "/", "Root Redirect", "Redirects to /admin (if logged in) or /login", "All users", "Yes"],
        ["PreRollTracker", "/login", "Login", "User authentication page", "All users", "No"],
        ["PreRollTracker", "/admin", "Admin Dashboard", "Main admin dashboard with batch management, live production tracking, and worker assignment", "Admins", "Yes"],
        ["PreRollTracker", "/admin/audit", "Audit Trail Viewer", "View timestamped audit trail of all system changes and user actions", "Admins", "Yes"],
        ["PreRollTracker", "/admin/inventory", "Inventory Management (Admin)", "Manage paper and cone inventory with stock levels and reorder alerts", "Admins", "Yes"],
        ["PreRollTracker", "/admin/settings", "System Settings", "Configure system-wide settings: backup schedule, notifications, feature flags", "Admins", "Yes"],
        ["PreRollTracker", "/archive", "Batch Archive", "Archive of completed production batches with search and historical data", "Workers", "Yes"],
        ["PreRollTracker", "/achievements", "Achievements Gallery", "Production milestone achievements and worker recognition badges", "Workers", "Yes"],
        ["PreRollTracker", "/finished-goods", "METRC Finished Goods", "METRC-integrated finished goods inventory tracking and reconciliation", "Workers / Admins", "Yes"],
        ["PreRollTracker", "/stats", "Production Statistics", "Charts and analytics for production metrics, trends, and performance KPIs", "Workers / Admins", "Yes"],
        ["PreRollTracker", "/wholesale", "Wholesale Inventory", "Mobile-friendly wholesale inventory view for field sales team", "Wholesale team", "Yes"],
        ["PreRollTracker", "/inventory", "Paper/Cone Inventory", "Worker-facing inventory management for papers, cones, and consumables", "Workers", "Yes"],
        ["PreRollTracker", "/centrifuge", "Centrifuge Calculator", "Interactive calculator for centrifuge parameters and mix ratios", "Workers", "Yes"],
        ["PreRollTracker", "/centrifuge-trends/<strain>", "Centrifuge Trends", "Historical centrifuge trend data and charts for a specific strain", "Workers", "Yes"],
        ["PreRollTracker", "/forgot-password", "Forgot Password", "Initiate password recovery via email link", "All users", "No"],
        ["PreRollTracker", "/reset-password", "Reset Password", "Password reset form (token-based from email link)", "All users", "No"],
        # ApexAPI screens
        ["ApexAPI", "Main Window", "Order Management", "Order search, filtering, status management, and label printing", "All users", "API token"],
        ["ApexAPI", "Pre-Roll Packing List", "Pre-Roll Packing List", "Two views: By Store and By Product with Google Sheets two-way sync", "Production", "API token"],
        ["ApexAPI", "Vape Cart Fill List", "Vape Cart Fill List", "Cart type categorization (510, Pod, AIO, Disposable) and fill planning", "Production", "API token"],
        ["ApexAPI", "Edibles Sell-Through", "Edibles Sell-Through", "Edibles sales tracking with production planning recommendations", "Management", "API token"],
        ["ApexAPI", "Inventory Management", "Inventory Management", "Manual inventory entry with autocomplete, recommendations, and quantity tracking", "Management", "API token"],
        ["ApexAPI", "Order Form Generator", "Order Form Generator", "CSV-to-PDF order form generation for wholesale accounts", "Admin", "API token"],
        ["ApexAPI", "Store Management", "Store Management", "Configure store aliases, delivery exclusions, and routing preferences", "Admin", "API token"],
    ]
    add_rows(ws, headers, data)
    style_sheet(ws, len(headers))


def build_api_endpoints_sheet(ws):
    headers = ["App", "Method", "Path", "Description", "Parameters", "Auth", "Response Type"]

    data = [
        # ── PreRollTracker: Auth ──
        ["PreRollTracker", "GET", "/login", "Render login page", "—", "None", "HTML"],
        ["PreRollTracker", "POST", "/login", "Authenticate user", "username, password", "None", "Redirect"],
        ["PreRollTracker", "GET", "/logout", "Log out current user", "—", "Session", "Redirect"],
        ["PreRollTracker", "GET", "/forgot-password", "Render forgot-password form", "—", "None", "HTML"],
        ["PreRollTracker", "POST", "/forgot-password", "Send password reset email", "email", "None", "JSON"],
        ["PreRollTracker", "GET", "/reset-password", "Render reset-password form", "token (query)", "None", "HTML"],
        ["PreRollTracker", "POST", "/reset-password", "Process password reset", "token, new_password", "None", "Redirect"],

        # ── PreRollTracker: Views ──
        ["PreRollTracker", "GET", "/", "Root redirect to /admin or /login", "—", "Session", "Redirect"],
        ["PreRollTracker", "GET", "/admin", "Admin dashboard", "—", "Admin session", "HTML"],
        ["PreRollTracker", "GET", "/admin/audit", "Audit trail page", "—", "Admin session", "HTML"],
        ["PreRollTracker", "GET", "/admin/inventory", "Admin inventory management", "—", "Admin session", "HTML"],
        ["PreRollTracker", "GET", "/admin/settings", "System settings page", "—", "Admin session", "HTML"],
        ["PreRollTracker", "GET", "/archive", "Batch archive page", "—", "Session", "HTML"],
        ["PreRollTracker", "GET", "/achievements", "Achievements gallery", "—", "Session", "HTML"],
        ["PreRollTracker", "GET", "/finished-goods", "Finished goods page", "—", "Session", "HTML"],
        ["PreRollTracker", "GET", "/stats", "Production statistics page", "—", "Session", "HTML"],
        ["PreRollTracker", "GET", "/wholesale", "Wholesale inventory view", "—", "Session", "HTML"],
        ["PreRollTracker", "GET", "/inventory", "Paper/cone inventory page", "—", "Session", "HTML"],
        ["PreRollTracker", "GET", "/centrifuge", "Centrifuge calculator page", "—", "Session", "HTML"],
        ["PreRollTracker", "GET", "/centrifuge-trends/<strain>", "Centrifuge trends for strain", "strain (path)", "Session", "HTML"],

        # ── PreRollTracker: API — Batches ──
        ["PreRollTracker", "GET", "/api/batches", "List all active batches", "status, limit, offset", "Session / API key", "JSON"],
        ["PreRollTracker", "POST", "/api/batches", "Create a new batch", "strain, size, target_qty, worker, etc.", "Admin session", "JSON"],
        ["PreRollTracker", "GET", "/api/batches/<id>", "Get batch details", "id (path)", "Session / API key", "JSON"],
        ["PreRollTracker", "PUT", "/api/batches/<id>", "Update batch fields", "id (path), fields to update", "Session / API key", "JSON"],
        ["PreRollTracker", "DELETE", "/api/batches/<id>", "Delete a batch", "id (path)", "Admin session", "JSON"],
        ["PreRollTracker", "POST", "/api/batches/<id>/start", "Start production on a batch", "id (path)", "Session", "JSON"],
        ["PreRollTracker", "POST", "/api/batches/<id>/pause", "Pause batch production", "id (path)", "Session", "JSON"],
        ["PreRollTracker", "POST", "/api/batches/<id>/resume", "Resume paused batch", "id (path)", "Session", "JSON"],
        ["PreRollTracker", "POST", "/api/batches/<id>/complete", "Mark batch complete", "id (path), final_count", "Session", "JSON"],
        ["PreRollTracker", "POST", "/api/batches/<id>/archive", "Archive a completed batch", "id (path)", "Admin session", "JSON"],
        ["PreRollTracker", "POST", "/api/batches/<id>/unarchive", "Unarchive a batch", "id (path)", "Admin session", "JSON"],
        ["PreRollTracker", "POST", "/api/batches/<id>/increment", "Increment batch count", "id (path), amount", "Session", "JSON"],
        ["PreRollTracker", "POST", "/api/batches/<id>/decrement", "Decrement batch count", "id (path), amount", "Session", "JSON"],
        ["PreRollTracker", "POST", "/api/batches/<id>/assign", "Assign worker to batch", "id (path), worker", "Admin session", "JSON"],
        ["PreRollTracker", "POST", "/api/batches/<id>/notes", "Add note to batch", "id (path), note", "Session", "JSON"],
        ["PreRollTracker", "GET", "/api/batches/<id>/history", "Get batch state history", "id (path)", "Session / API key", "JSON"],
        ["PreRollTracker", "POST", "/api/batches/<id>/qc", "Record QC check result", "id (path), passed, notes", "Session", "JSON"],
        ["PreRollTracker", "POST", "/api/batches/reorder", "Reorder batch display positions", "order (array of ids)", "Admin session", "JSON"],
        ["PreRollTracker", "GET", "/api/batches/active", "Get only active batches", "—", "Session / API key", "JSON"],
        ["PreRollTracker", "GET", "/api/batches/archived", "Get only archived batches", "limit, offset, search", "Session / API key", "JSON"],
        ["PreRollTracker", "GET", "/api/batches/stats", "Aggregate batch statistics", "date_from, date_to", "Session / API key", "JSON"],
        ["PreRollTracker", "POST", "/api/batches/<id>/duplicate", "Duplicate a batch", "id (path)", "Admin session", "JSON"],

        # ── PreRollTracker: API — Inventory ──
        ["PreRollTracker", "GET", "/api/inventory", "List all inventory items", "category", "Session / API key", "JSON"],
        ["PreRollTracker", "POST", "/api/inventory", "Add inventory item", "name, category, quantity, unit", "Session", "JSON"],
        ["PreRollTracker", "PUT", "/api/inventory/<id>", "Update inventory item", "id (path), fields", "Session", "JSON"],
        ["PreRollTracker", "DELETE", "/api/inventory/<id>", "Delete inventory item", "id (path)", "Admin session", "JSON"],
        ["PreRollTracker", "POST", "/api/inventory/<id>/adjust", "Adjust item quantity", "id (path), adjustment, reason", "Session", "JSON"],
        ["PreRollTracker", "GET", "/api/inventory/low-stock", "Get items below reorder point", "—", "Session / API key", "JSON"],
        ["PreRollTracker", "GET", "/api/inventory/usage", "Get inventory usage history", "item_id, date_from, date_to", "Session / API key", "JSON"],
        ["PreRollTracker", "POST", "/api/inventory/bulk-update", "Bulk update inventory quantities", "items (array)", "Admin session", "JSON"],

        # ── PreRollTracker: API — Finished Goods ──
        ["PreRollTracker", "GET", "/api/finished-goods", "List finished goods inventory", "search, category, sort", "Session / API key", "JSON"],
        ["PreRollTracker", "POST", "/api/finished-goods", "Add finished goods entry", "sku, strain, size, quantity, metrc_tag", "Session", "JSON"],
        ["PreRollTracker", "PUT", "/api/finished-goods/<id>", "Update finished goods entry", "id (path), fields", "Session", "JSON"],
        ["PreRollTracker", "DELETE", "/api/finished-goods/<id>", "Delete finished goods entry", "id (path)", "Admin session", "JSON"],
        ["PreRollTracker", "POST", "/api/finished-goods/<id>/adjust", "Adjust FG quantity", "id (path), adjustment, reason", "Session", "JSON"],
        ["PreRollTracker", "GET", "/api/finished-goods/history", "Finished goods change history", "item_id, limit", "Session / API key", "JSON"],
        ["PreRollTracker", "GET", "/api/finished-goods/summary", "Aggregated FG summary", "group_by", "Session / API key", "JSON"],
        ["PreRollTracker", "POST", "/api/finished-goods/reconcile", "Reconcile FG with METRC", "items (array)", "Admin session", "JSON"],
        ["PreRollTracker", "GET", "/api/finished-goods/gram-tracking", "Gram weight tracking data", "strain, date_from, date_to", "Session / API key", "JSON"],

        # ── PreRollTracker: API — Wholesale ──
        ["PreRollTracker", "GET", "/api/wholesale", "List wholesale inventory", "search, available_only", "Session / API key", "JSON"],
        ["PreRollTracker", "POST", "/api/wholesale/hold", "Place wholesale hold", "item_id, quantity, customer, notes", "Session", "JSON"],
        ["PreRollTracker", "DELETE", "/api/wholesale/hold/<id>", "Release wholesale hold", "id (path)", "Session", "JSON"],
        ["PreRollTracker", "POST", "/api/wholesale/hold/<id>/confirm", "Confirm hold as sold", "id (path)", "Session", "JSON"],
        ["PreRollTracker", "GET", "/api/wholesale/holds", "List all active holds", "customer", "Session / API key", "JSON"],
        ["PreRollTracker", "GET", "/api/wholesale/availability", "Real-time availability check", "item_ids (array)", "Session / API key", "JSON"],

        # ── PreRollTracker: API — Centrifuge ──
        ["PreRollTracker", "GET", "/api/centrifuge/strains", "List strains with centrifuge data", "—", "Session / API key", "JSON"],
        ["PreRollTracker", "POST", "/api/centrifuge/calculate", "Run centrifuge calculation", "strain, weight, moisture", "Session", "JSON"],
        ["PreRollTracker", "GET", "/api/centrifuge/trends/<strain>", "Get centrifuge trends", "strain (path), days", "Session / API key", "JSON"],
        ["PreRollTracker", "POST", "/api/centrifuge/log", "Log centrifuge run result", "strain, params, result", "Session", "JSON"],

        # ── PreRollTracker: API — Snapshots ──
        ["PreRollTracker", "GET", "/api/snapshots", "List production snapshots", "date_from, date_to", "Session / API key", "JSON"],
        ["PreRollTracker", "POST", "/api/snapshots", "Create manual snapshot", "—", "Admin session", "JSON"],
        ["PreRollTracker", "GET", "/api/snapshots/<id>", "Get snapshot details", "id (path)", "Session / API key", "JSON"],
        ["PreRollTracker", "GET", "/api/snapshots/latest", "Get most recent snapshot", "—", "Session / API key", "JSON"],
        ["PreRollTracker", "GET", "/api/snapshots/compare", "Compare two snapshots", "snapshot_a, snapshot_b", "Session / API key", "JSON"],

        # ── PreRollTracker: API — Misc ──
        ["PreRollTracker", "GET", "/api/audit", "Get audit log entries", "action, user, date_from, date_to, limit", "Admin session", "JSON"],
        ["PreRollTracker", "GET", "/api/settings", "Get all settings", "—", "Admin session", "JSON"],
        ["PreRollTracker", "PUT", "/api/settings", "Update settings", "key-value pairs", "Admin session", "JSON"],
        ["PreRollTracker", "GET", "/api/settings/<key>", "Get single setting value", "key (path)", "Admin session", "JSON"],
        ["PreRollTracker", "GET", "/api/health", "Health check endpoint", "—", "None", "JSON"],
        ["PreRollTracker", "GET", "/api/version", "Application version info", "—", "None", "JSON"],
        ["PreRollTracker", "POST", "/api/backup/trigger", "Trigger manual backup", "—", "Admin session", "JSON"],
        ["PreRollTracker", "GET", "/api/backup/status", "Get backup status", "—", "Admin session", "JSON"],
        ["PreRollTracker", "GET", "/api/learning/rates", "Get learned production rates", "strain, size", "Session / API key", "JSON"],
        ["PreRollTracker", "POST", "/api/learning/reset", "Reset learning data", "strain, size", "Admin session", "JSON"],
        ["PreRollTracker", "GET", "/api/workers", "List all workers", "active_only", "Session / API key", "JSON"],
        ["PreRollTracker", "POST", "/api/workers", "Add a new worker", "name, role", "Admin session", "JSON"],
        ["PreRollTracker", "PUT", "/api/workers/<id>", "Update worker info", "id (path), fields", "Admin session", "JSON"],
        ["PreRollTracker", "DELETE", "/api/workers/<id>", "Deactivate worker", "id (path)", "Admin session", "JSON"],

        # ── PreRollTracker: PWA ──
        ["PreRollTracker", "GET", "/manifest.json", "PWA manifest file", "—", "None", "JSON"],
        ["PreRollTracker", "GET", "/sw.js", "Service worker script", "—", "None", "JS"],
        ["PreRollTracker", "GET", "/offline", "Offline fallback page", "—", "None", "HTML"],

        # ── ApexAPI: Apex Trading API (External) ──
        ["ApexAPI", "GET", "/api/v1/orders", "Fetch orders from Apex Trading", "status, date_from, date_to, page", "Bearer token", "JSON"],
        ["ApexAPI", "GET", "/api/v1/orders/<id>", "Get single order details", "id (path)", "Bearer token", "JSON"],
        ["ApexAPI", "PUT", "/api/v1/orders/<id>/status", "Update order status in Apex", "id (path), status", "Bearer token", "JSON"],
        ["ApexAPI", "GET", "/api/v2/batches", "List product batches (v2)", "product_id, status", "Bearer token", "JSON"],
        ["ApexAPI", "GET", "/api/v2/batches/<id>", "Get batch details (v2)", "id (path)", "Bearer token", "JSON"],
        ["ApexAPI", "PUT", "/api/v2/batches/<id>/inventory", "Update batch inventory quantity", "id (path), quantity", "Bearer token", "JSON"],
        ["ApexAPI", "GET", "/api/v1/products", "List all products", "category, active", "Bearer token", "JSON"],
        ["ApexAPI", "GET", "/api/v1/products/<id>", "Get product details", "id (path)", "Bearer token", "JSON"],
        ["ApexAPI", "GET", "/api/v1/buyers", "List all buyer accounts", "search", "Bearer token", "JSON"],
        ["ApexAPI", "GET", "/api/v1/buyers/<id>", "Get buyer details", "id (path)", "Bearer token", "JSON"],
        ["ApexAPI", "GET", "/api/v1/cannabinoids/<batch_id>", "Get cannabinoid test results", "batch_id (path)", "Bearer token", "JSON"],

        # ── ApexAPI: Dashboard API (PreRollTracker) ──
        ["ApexAPI", "GET", "/api/finished-goods (Dashboard)", "Fetch finished goods from PreRollTracker dashboard", "—", "API key", "JSON"],
        ["ApexAPI", "GET", "/api/finished-goods/gram-tracking (Dashboard)", "Fetch gram tracking data from dashboard", "strain", "API key", "JSON"],
        ["ApexAPI", "GET", "/api/batches/active (Dashboard)", "Fetch active production batches", "—", "API key", "JSON"],
    ]

    add_rows(ws, headers, data)
    style_sheet(ws, len(headers))


def build_database_tables_sheet(ws):
    headers = ["App", "Table Name", "Column", "Type", "Constraints", "Description"]

    data = [
        # ── batches ──
        ["PreRollTracker", "batches", "id", "INTEGER", "PK, AUTOINCREMENT", "Unique batch identifier"],
        ["PreRollTracker", "batches", "strain", "TEXT", "NOT NULL", "Cannabis strain name"],
        ["PreRollTracker", "batches", "size", "TEXT", "NOT NULL", "Pre-roll size (0.5g, 1g, 1.5g, etc.)"],
        ["PreRollTracker", "batches", "target_quantity", "INTEGER", "NOT NULL", "Target production quantity"],
        ["PreRollTracker", "batches", "current_count", "INTEGER", "DEFAULT 0", "Current completed count"],
        ["PreRollTracker", "batches", "status", "TEXT", "DEFAULT 'pending'", "Batch status: pending, active, paused, completed, archived"],
        ["PreRollTracker", "batches", "worker", "TEXT", "", "Assigned worker name"],
        ["PreRollTracker", "batches", "created_at", "TIMESTAMP", "DEFAULT CURRENT_TIMESTAMP", "Batch creation timestamp"],
        ["PreRollTracker", "batches", "started_at", "TIMESTAMP", "", "Production start time"],
        ["PreRollTracker", "batches", "completed_at", "TIMESTAMP", "", "Production completion time"],
        ["PreRollTracker", "batches", "archived_at", "TIMESTAMP", "", "Archive timestamp"],
        ["PreRollTracker", "batches", "updated_at", "TIMESTAMP", "", "Last update timestamp"],
        ["PreRollTracker", "batches", "notes", "TEXT", "", "Free-text batch notes"],
        ["PreRollTracker", "batches", "priority", "INTEGER", "DEFAULT 0", "Display priority / sort order"],
        ["PreRollTracker", "batches", "display_order", "INTEGER", "", "Manual display ordering"],
        ["PreRollTracker", "batches", "paper_type", "TEXT", "", "Paper/cone type used"],
        ["PreRollTracker", "batches", "filter_type", "TEXT", "", "Filter type used"],
        ["PreRollTracker", "batches", "machine", "TEXT", "", "Machine/station used"],
        ["PreRollTracker", "batches", "metrc_tag", "TEXT", "", "METRC package tag"],
        ["PreRollTracker", "batches", "metrc_source_tag", "TEXT", "", "METRC source package tag"],
        ["PreRollTracker", "batches", "lot_number", "TEXT", "", "Internal lot/batch number"],
        ["PreRollTracker", "batches", "harvest_date", "DATE", "", "Harvest date of source material"],
        ["PreRollTracker", "batches", "package_date", "DATE", "", "Package date for METRC"],
        ["PreRollTracker", "batches", "test_results", "TEXT", "", "JSON-encoded cannabinoid test results"],
        ["PreRollTracker", "batches", "thc_percentage", "REAL", "", "THC percentage from testing"],
        ["PreRollTracker", "batches", "cbd_percentage", "REAL", "", "CBD percentage from testing"],
        ["PreRollTracker", "batches", "total_terpenes", "REAL", "", "Total terpene percentage"],
        ["PreRollTracker", "batches", "moisture_content", "REAL", "", "Moisture content percentage"],
        ["PreRollTracker", "batches", "target_weight", "REAL", "", "Target weight per pre-roll in grams"],
        ["PreRollTracker", "batches", "actual_weight_avg", "REAL", "", "Average actual weight per pre-roll"],
        ["PreRollTracker", "batches", "weight_variance", "REAL", "", "Weight variance from target"],
        ["PreRollTracker", "batches", "material_weight_start", "REAL", "", "Starting material weight (grams)"],
        ["PreRollTracker", "batches", "material_weight_end", "REAL", "", "Ending material weight (grams)"],
        ["PreRollTracker", "batches", "material_waste", "REAL", "", "Material waste (grams)"],
        ["PreRollTracker", "batches", "qc_passed", "BOOLEAN", "", "QC inspection pass/fail"],
        ["PreRollTracker", "batches", "qc_notes", "TEXT", "", "QC inspection notes"],
        ["PreRollTracker", "batches", "qc_checked_at", "TIMESTAMP", "", "QC check timestamp"],
        ["PreRollTracker", "batches", "qc_checked_by", "TEXT", "", "QC inspector name"],
        ["PreRollTracker", "batches", "defect_count", "INTEGER", "DEFAULT 0", "Number of defective units"],
        ["PreRollTracker", "batches", "defect_types", "TEXT", "", "JSON array of defect type codes"],
        ["PreRollTracker", "batches", "production_rate", "REAL", "", "Calculated production rate (units/hr)"],
        ["PreRollTracker", "batches", "estimated_completion", "TIMESTAMP", "", "ML-predicted completion time"],
        ["PreRollTracker", "batches", "time_active_seconds", "INTEGER", "DEFAULT 0", "Total active production seconds"],
        ["PreRollTracker", "batches", "time_paused_seconds", "INTEGER", "DEFAULT 0", "Total paused seconds"],
        ["PreRollTracker", "batches", "pause_count", "INTEGER", "DEFAULT 0", "Number of times paused"],
        ["PreRollTracker", "batches", "last_increment_at", "TIMESTAMP", "", "Timestamp of last count increment"],
        ["PreRollTracker", "batches", "increment_history", "TEXT", "", "JSON array of increment events"],
        ["PreRollTracker", "batches", "centrifuge_speed", "INTEGER", "", "Centrifuge speed (RPM)"],
        ["PreRollTracker", "batches", "centrifuge_time", "INTEGER", "", "Centrifuge duration (seconds)"],
        ["PreRollTracker", "batches", "centrifuge_temp", "REAL", "", "Centrifuge temperature"],
        ["PreRollTracker", "batches", "humidity", "REAL", "", "Ambient humidity during production"],
        ["PreRollTracker", "batches", "temperature", "REAL", "", "Ambient temperature during production"],
        ["PreRollTracker", "batches", "room", "TEXT", "", "Production room identifier"],
        ["PreRollTracker", "batches", "shift", "TEXT", "", "Production shift (day/swing/night)"],
        ["PreRollTracker", "batches", "category", "TEXT", "", "Product category classification"],
        ["PreRollTracker", "batches", "brand", "TEXT", "", "Brand name"],
        ["PreRollTracker", "batches", "sku", "TEXT", "", "Product SKU"],
        ["PreRollTracker", "batches", "upc", "TEXT", "", "UPC barcode"],
        ["PreRollTracker", "batches", "unit_price", "REAL", "", "Unit wholesale price"],
        ["PreRollTracker", "batches", "total_value", "REAL", "", "Total batch value"],
        ["PreRollTracker", "batches", "destination", "TEXT", "", "Intended destination/customer"],
        ["PreRollTracker", "batches", "is_learning_processed", "BOOLEAN", "DEFAULT 0", "Whether learning processor has analyzed this batch"],

        # ── audit_log ──
        ["PreRollTracker", "audit_log", "id", "INTEGER", "PK, AUTOINCREMENT", "Unique log entry ID"],
        ["PreRollTracker", "audit_log", "timestamp", "TIMESTAMP", "DEFAULT CURRENT_TIMESTAMP", "Event timestamp"],
        ["PreRollTracker", "audit_log", "user", "TEXT", "", "User who performed the action"],
        ["PreRollTracker", "audit_log", "action", "TEXT", "NOT NULL", "Action type (create, update, delete, login, etc.)"],
        ["PreRollTracker", "audit_log", "entity_type", "TEXT", "", "Entity type affected (batch, inventory, setting, etc.)"],
        ["PreRollTracker", "audit_log", "entity_id", "INTEGER", "", "ID of affected entity"],
        ["PreRollTracker", "audit_log", "details", "TEXT", "", "JSON-encoded change details (old/new values)"],
        ["PreRollTracker", "audit_log", "ip_address", "TEXT", "", "Client IP address"],

        # ── finished_goods ──
        ["PreRollTracker", "finished_goods", "id", "INTEGER", "PK, AUTOINCREMENT", "Unique finished goods ID"],
        ["PreRollTracker", "finished_goods", "sku", "TEXT", "", "Product SKU"],
        ["PreRollTracker", "finished_goods", "strain", "TEXT", "NOT NULL", "Cannabis strain name"],
        ["PreRollTracker", "finished_goods", "size", "TEXT", "NOT NULL", "Pre-roll size"],
        ["PreRollTracker", "finished_goods", "quantity", "INTEGER", "NOT NULL", "Current quantity on hand"],
        ["PreRollTracker", "finished_goods", "metrc_tag", "TEXT", "", "METRC package tag"],
        ["PreRollTracker", "finished_goods", "category", "TEXT", "", "Product category"],
        ["PreRollTracker", "finished_goods", "brand", "TEXT", "", "Brand name"],
        ["PreRollTracker", "finished_goods", "created_at", "TIMESTAMP", "DEFAULT CURRENT_TIMESTAMP", "Record creation time"],
        ["PreRollTracker", "finished_goods", "updated_at", "TIMESTAMP", "", "Last update time"],
        ["PreRollTracker", "finished_goods", "batch_id", "INTEGER", "FK → batches.id", "Source batch reference"],
        ["PreRollTracker", "finished_goods", "unit_weight", "REAL", "", "Weight per unit in grams"],
        ["PreRollTracker", "finished_goods", "total_weight", "REAL", "", "Total weight in grams"],

        # ── finished_goods_history ──
        ["PreRollTracker", "finished_goods_history", "id", "INTEGER", "PK, AUTOINCREMENT", "History entry ID"],
        ["PreRollTracker", "finished_goods_history", "finished_goods_id", "INTEGER", "FK → finished_goods.id", "Referenced finished goods item"],
        ["PreRollTracker", "finished_goods_history", "action", "TEXT", "NOT NULL", "Action (add, adjust, reconcile)"],
        ["PreRollTracker", "finished_goods_history", "quantity_change", "INTEGER", "", "Quantity delta (+/-)"],
        ["PreRollTracker", "finished_goods_history", "quantity_before", "INTEGER", "", "Quantity before change"],
        ["PreRollTracker", "finished_goods_history", "quantity_after", "INTEGER", "", "Quantity after change"],
        ["PreRollTracker", "finished_goods_history", "reason", "TEXT", "", "Reason for adjustment"],
        ["PreRollTracker", "finished_goods_history", "user", "TEXT", "", "User who made the change"],
        ["PreRollTracker", "finished_goods_history", "timestamp", "TIMESTAMP", "DEFAULT CURRENT_TIMESTAMP", "Change timestamp"],

        # ── inventory ──
        ["PreRollTracker", "inventory", "id", "INTEGER", "PK, AUTOINCREMENT", "Inventory item ID"],
        ["PreRollTracker", "inventory", "name", "TEXT", "NOT NULL, UNIQUE", "Item name"],
        ["PreRollTracker", "inventory", "category", "TEXT", "NOT NULL", "Category (paper, cone, filter, tube, label, etc.)"],
        ["PreRollTracker", "inventory", "quantity", "REAL", "DEFAULT 0", "Current quantity on hand"],
        ["PreRollTracker", "inventory", "unit", "TEXT", "", "Unit of measure"],
        ["PreRollTracker", "inventory", "reorder_point", "REAL", "", "Low-stock reorder threshold"],
        ["PreRollTracker", "inventory", "reorder_quantity", "REAL", "", "Suggested reorder quantity"],
        ["PreRollTracker", "inventory", "supplier", "TEXT", "", "Primary supplier name"],
        ["PreRollTracker", "inventory", "notes", "TEXT", "", "Item notes"],
        ["PreRollTracker", "inventory", "created_at", "TIMESTAMP", "DEFAULT CURRENT_TIMESTAMP", "Record created"],
        ["PreRollTracker", "inventory", "updated_at", "TIMESTAMP", "", "Last updated"],

        # ── inventory_usage ──
        ["PreRollTracker", "inventory_usage", "id", "INTEGER", "PK, AUTOINCREMENT", "Usage record ID"],
        ["PreRollTracker", "inventory_usage", "inventory_id", "INTEGER", "FK → inventory.id", "Referenced inventory item"],
        ["PreRollTracker", "inventory_usage", "batch_id", "INTEGER", "FK → batches.id", "Associated batch (if any)"],
        ["PreRollTracker", "inventory_usage", "quantity_used", "REAL", "NOT NULL", "Quantity consumed"],
        ["PreRollTracker", "inventory_usage", "timestamp", "TIMESTAMP", "DEFAULT CURRENT_TIMESTAMP", "Usage timestamp"],
        ["PreRollTracker", "inventory_usage", "user", "TEXT", "", "User who recorded usage"],

        # ── settings ──
        ["PreRollTracker", "settings", "key", "TEXT", "PK", "Setting key name"],
        ["PreRollTracker", "settings", "value", "TEXT", "", "Setting value (JSON-encoded for complex types)"],
        ["PreRollTracker", "settings", "updated_at", "TIMESTAMP", "DEFAULT CURRENT_TIMESTAMP", "Last modified"],
        ["PreRollTracker", "settings", "updated_by", "TEXT", "", "User who last changed this setting"],

        # ── wholesale_holds ──
        ["PreRollTracker", "wholesale_holds", "id", "INTEGER", "PK, AUTOINCREMENT", "Hold record ID"],
        ["PreRollTracker", "wholesale_holds", "finished_goods_id", "INTEGER", "FK → finished_goods.id", "Item being held"],
        ["PreRollTracker", "wholesale_holds", "quantity", "INTEGER", "NOT NULL", "Quantity on hold"],
        ["PreRollTracker", "wholesale_holds", "customer", "TEXT", "NOT NULL", "Customer/account name"],
        ["PreRollTracker", "wholesale_holds", "notes", "TEXT", "", "Hold notes"],
        ["PreRollTracker", "wholesale_holds", "status", "TEXT", "DEFAULT 'active'", "Hold status: active, confirmed, released"],
        ["PreRollTracker", "wholesale_holds", "created_at", "TIMESTAMP", "DEFAULT CURRENT_TIMESTAMP", "Hold placed timestamp"],
        ["PreRollTracker", "wholesale_holds", "confirmed_at", "TIMESTAMP", "", "Hold confirmed as sold timestamp"],
        ["PreRollTracker", "wholesale_holds", "released_at", "TIMESTAMP", "", "Hold released timestamp"],
        ["PreRollTracker", "wholesale_holds", "created_by", "TEXT", "", "User who placed the hold"],

        # ── schema_version ──
        ["PreRollTracker", "schema_version", "version", "INTEGER", "PK", "Schema version number"],
        ["PreRollTracker", "schema_version", "applied_at", "TIMESTAMP", "DEFAULT CURRENT_TIMESTAMP", "Migration applied timestamp"],
        ["PreRollTracker", "schema_version", "description", "TEXT", "", "Migration description"],

        # ═══ ApexAPI Tables ═══

        # ── orders ──
        ["ApexAPI", "orders", "id", "TEXT", "PK", "Apex Trading order ID"],
        ["ApexAPI", "orders", "order_number", "TEXT", "", "Human-readable order number"],
        ["ApexAPI", "orders", "status", "TEXT", "", "Order status (pending, confirmed, shipped, etc.)"],
        ["ApexAPI", "orders", "buyer_name", "TEXT", "", "Buyer/account name"],
        ["ApexAPI", "orders", "buyer_id", "TEXT", "", "Apex buyer account ID"],
        ["ApexAPI", "orders", "total", "REAL", "", "Order total amount"],
        ["ApexAPI", "orders", "items", "TEXT", "", "JSON-encoded line items"],
        ["ApexAPI", "orders", "created_at", "TIMESTAMP", "", "Order creation date"],
        ["ApexAPI", "orders", "updated_at", "TIMESTAMP", "", "Last update from Apex"],
        ["ApexAPI", "orders", "delivery_date", "DATE", "", "Scheduled delivery date"],
        ["ApexAPI", "orders", "notes", "TEXT", "", "Order notes"],
        ["ApexAPI", "orders", "cached_at", "TIMESTAMP", "", "When this record was cached locally"],

        # ── print_status ──
        ["ApexAPI", "print_status", "id", "INTEGER", "PK, AUTOINCREMENT", "Print record ID"],
        ["ApexAPI", "print_status", "order_id", "TEXT", "FK → orders.id", "Associated order"],
        ["ApexAPI", "print_status", "printed_at", "TIMESTAMP", "DEFAULT CURRENT_TIMESTAMP", "Print timestamp"],
        ["ApexAPI", "print_status", "printer_name", "TEXT", "", "Printer used"],
        ["ApexAPI", "print_status", "copies", "INTEGER", "DEFAULT 1", "Number of copies printed"],
        ["ApexAPI", "print_status", "status", "TEXT", "", "Print job status"],

        # ── order_cache_metadata ──
        ["ApexAPI", "order_cache_metadata", "key", "TEXT", "PK", "Cache key identifier"],
        ["ApexAPI", "order_cache_metadata", "value", "TEXT", "", "Cached value"],
        ["ApexAPI", "order_cache_metadata", "expires_at", "TIMESTAMP", "", "Cache expiration time"],
        ["ApexAPI", "order_cache_metadata", "created_at", "TIMESTAMP", "DEFAULT CURRENT_TIMESTAMP", "Cache entry creation"],

        # ── order_status_history ──
        ["ApexAPI", "order_status_history", "id", "INTEGER", "PK, AUTOINCREMENT", "History entry ID"],
        ["ApexAPI", "order_status_history", "order_id", "TEXT", "FK → orders.id", "Associated order"],
        ["ApexAPI", "order_status_history", "old_status", "TEXT", "", "Previous status"],
        ["ApexAPI", "order_status_history", "new_status", "TEXT", "", "New status"],
        ["ApexAPI", "order_status_history", "changed_at", "TIMESTAMP", "DEFAULT CURRENT_TIMESTAMP", "Status change timestamp"],
        ["ApexAPI", "order_status_history", "changed_by", "TEXT", "", "User or system that changed status"],

        # ── config ──
        ["ApexAPI", "config", "key", "TEXT", "PK", "Configuration key"],
        ["ApexAPI", "config", "value", "TEXT", "", "Configuration value"],
        ["ApexAPI", "config", "updated_at", "TIMESTAMP", "", "Last modified"],

        # ── user_preferences ──
        ["ApexAPI", "user_preferences", "key", "TEXT", "PK", "Preference key"],
        ["ApexAPI", "user_preferences", "value", "TEXT", "", "Preference value (JSON)"],
        ["ApexAPI", "user_preferences", "updated_at", "TIMESTAMP", "", "Last modified"],

        # ── config_history ──
        ["ApexAPI", "config_history", "id", "INTEGER", "PK, AUTOINCREMENT", "History entry ID"],
        ["ApexAPI", "config_history", "key", "TEXT", "NOT NULL", "Config key that changed"],
        ["ApexAPI", "config_history", "old_value", "TEXT", "", "Previous value"],
        ["ApexAPI", "config_history", "new_value", "TEXT", "", "New value"],
        ["ApexAPI", "config_history", "changed_at", "TIMESTAMP", "DEFAULT CURRENT_TIMESTAMP", "Change timestamp"],

        # ── system_settings ──
        ["ApexAPI", "system_settings", "key", "TEXT", "PK", "System setting key"],
        ["ApexAPI", "system_settings", "value", "TEXT", "", "Setting value"],
        ["ApexAPI", "system_settings", "description", "TEXT", "", "Human-readable description"],
        ["ApexAPI", "system_settings", "updated_at", "TIMESTAMP", "", "Last modified"],

        # ── preroll_items ──
        ["ApexAPI", "preroll_items", "id", "INTEGER", "PK, AUTOINCREMENT", "Pre-roll item ID"],
        ["ApexAPI", "preroll_items", "product_name", "TEXT", "NOT NULL", "Product name from Apex"],
        ["ApexAPI", "preroll_items", "strain", "TEXT", "", "Strain name"],
        ["ApexAPI", "preroll_items", "size", "TEXT", "", "Pre-roll size"],
        ["ApexAPI", "preroll_items", "quantity", "INTEGER", "", "Quantity needed"],
        ["ApexAPI", "preroll_items", "store_name", "TEXT", "", "Destination store"],
        ["ApexAPI", "preroll_items", "order_id", "TEXT", "FK → orders.id", "Source order"],
        ["ApexAPI", "preroll_items", "batch_id", "TEXT", "", "Apex batch ID"],
        ["ApexAPI", "preroll_items", "status", "TEXT", "DEFAULT 'pending'", "Packing status"],
        ["ApexAPI", "preroll_items", "created_at", "TIMESTAMP", "DEFAULT CURRENT_TIMESTAMP", "Record created"],

        # ── preroll_metadata ──
        ["ApexAPI", "preroll_metadata", "id", "INTEGER", "PK, AUTOINCREMENT", "Metadata entry ID"],
        ["ApexAPI", "preroll_metadata", "sync_date", "TIMESTAMP", "", "Last sync with Google Sheets"],
        ["ApexAPI", "preroll_metadata", "total_items", "INTEGER", "", "Total items in packing list"],
        ["ApexAPI", "preroll_metadata", "total_stores", "INTEGER", "", "Number of stores"],
        ["ApexAPI", "preroll_metadata", "status", "TEXT", "", "Overall sync status"],

        # ── preroll_items_archive ──
        ["ApexAPI", "preroll_items_archive", "id", "INTEGER", "PK, AUTOINCREMENT", "Archive entry ID"],
        ["ApexAPI", "preroll_items_archive", "original_id", "INTEGER", "", "Original preroll_items.id"],
        ["ApexAPI", "preroll_items_archive", "product_name", "TEXT", "", "Product name"],
        ["ApexAPI", "preroll_items_archive", "strain", "TEXT", "", "Strain name"],
        ["ApexAPI", "preroll_items_archive", "quantity", "INTEGER", "", "Quantity"],
        ["ApexAPI", "preroll_items_archive", "store_name", "TEXT", "", "Destination store"],
        ["ApexAPI", "preroll_items_archive", "archived_at", "TIMESTAMP", "DEFAULT CURRENT_TIMESTAMP", "Archive timestamp"],

        # ── edibles_orders ──
        ["ApexAPI", "edibles_orders", "id", "INTEGER", "PK, AUTOINCREMENT", "Edibles order ID"],
        ["ApexAPI", "edibles_orders", "product_name", "TEXT", "NOT NULL", "Edible product name"],
        ["ApexAPI", "edibles_orders", "quantity_ordered", "INTEGER", "", "Quantity ordered"],
        ["ApexAPI", "edibles_orders", "quantity_sold", "INTEGER", "", "Quantity sold through"],
        ["ApexAPI", "edibles_orders", "store_name", "TEXT", "", "Store name"],
        ["ApexAPI", "edibles_orders", "order_date", "DATE", "", "Order date"],
        ["ApexAPI", "edibles_orders", "sell_through_rate", "REAL", "", "Calculated sell-through percentage"],
        ["ApexAPI", "edibles_orders", "created_at", "TIMESTAMP", "DEFAULT CURRENT_TIMESTAMP", "Record created"],

        # ── migrations ──
        ["ApexAPI", "migrations", "id", "INTEGER", "PK, AUTOINCREMENT", "Migration record ID"],
        ["ApexAPI", "migrations", "version", "INTEGER", "NOT NULL, UNIQUE", "Migration version number"],
        ["ApexAPI", "migrations", "description", "TEXT", "", "Migration description"],
        ["ApexAPI", "migrations", "applied_at", "TIMESTAMP", "DEFAULT CURRENT_TIMESTAMP", "Applied timestamp"],
    ]

    add_rows(ws, headers, data)
    style_sheet(ws, len(headers))


def build_background_jobs_sheet(ws):
    headers = ["App", "Job Name", "Schedule", "Description", "Configuration"]

    data = [
        [
            "PreRollTracker",
            "Learning Processor",
            "Every 5 minutes",
            "Background thread that processes completed batch data to learn production rates per strain/size combination. Uses weighted moving average for predictions. Runs as a daemon thread started on app init.",
            "Automatic; no user config. Processes batches where is_learning_processed = 0. Thread started in app factory.",
        ],
        [
            "PreRollTracker",
            "Backup Scheduler",
            "Every 6 hours",
            "AES-256 encrypted database backups. Compresses SQLite DB, encrypts with Fernet, uploads to local backup dir and GitHub Releases. Uses worker election lock to prevent duplicate runs in multi-worker deployments.",
            "BACKUP_ENCRYPTION_KEY (env), GITHUB_TOKEN (env), GITHUB_REPO (env), PUSHOVER_APP_TOKEN + PUSHOVER_USER_KEY (env, for failure alerts). Configurable schedule via settings table.",
        ],
        [
            "ApexAPI",
            "Cache Warmer",
            "Configurable (2min to 24hr per strategy)",
            "Proactive cache preloading system with multiple configurable strategies (orders, products, batches). Prevents cold-cache latency for frequently accessed data. Supports priority-based warming order.",
            "cache_warming.enabled (bool), cache_warming.strategies (dict with intervals per data type), cache_warming.max_concurrent (int). Set in apex_config.json.",
        ],
        [
            "ApexAPI",
            "Batch Inventory Sync",
            "Every 5 seconds",
            "Automatically synchronizes local inventory quantity changes back to the Apex Trading API. Detects modified batch quantities and pushes updates. Includes retry logic and conflict resolution.",
            "batch_inventory_sync.enabled (bool), batch_inventory_sync.interval_seconds (int, default 5), batch_inventory_sync.retry_count (int). Set in apex_config.json.",
        ],
        [
            "ApexAPI",
            "Auto-Refresh",
            "Every 5 seconds",
            "Periodic automatic refresh of the main order list in the GUI. Fetches latest order data from Apex Trading API and updates the display. Can be toggled on/off via the UI.",
            "auto_refresh_enabled (bool), auto_refresh_interval (int, seconds). Toggled in GUI toolbar. Persisted in user_preferences table.",
        ],
    ]

    add_rows(ws, headers, data)
    style_sheet(ws, len(headers))


def build_integrations_sheet(ws):
    headers = ["Source App", "Target", "Integration Type", "Data Flow", "Endpoint / Method"]

    data = [
        [
            "ApexAPI",
            "Apex Trading API",
            "REST API (Bearer token)",
            "Bidirectional: reads orders, batches, products, buyers, cannabinoids; writes order status updates and inventory quantities",
            "https://app.apextrading.com/api/v1/* and /api/v2/*. Auth via Bearer token in Authorization header.",
        ],
        [
            "ApexAPI",
            "PreRollTracker Dashboard",
            "REST API (API key)",
            "Read-only: fetches finished goods inventory, gram tracking data, and active batch status for display in desktop GUI",
            "https://himomstats.online/api/finished-goods, /api/finished-goods/gram-tracking, /api/batches/active. Auth via X-API-Key header.",
        ],
        [
            "ApexAPI",
            "Google Sheets",
            "OAuth 2.0 (Google Sheets API v4)",
            "Bidirectional: pushes pre-roll packing list data to shared spreadsheet; pulls manual edits back. Supports named ranges and batch updates.",
            "Google Sheets API v4 via gspread library. OAuth 2.0 credentials stored locally. Sheet ID configured in apex_config.json.",
        ],
        [
            "PreRollTracker",
            "Sentry",
            "SDK (sentry-sdk)",
            "Outbound: sends unhandled exceptions, performance traces, and breadcrumbs for error monitoring and alerting",
            "Sentry Python SDK (Flask integration). Configured via SENTRY_DSN env var. Environment tag from SENTRY_ENV.",
        ],
        [
            "PreRollTracker",
            "Pushover",
            "REST API (HTTPS POST)",
            "Outbound: sends push notifications to admin devices on backup failures, critical errors, and low-stock alerts",
            "https://api.pushover.net/1/messages.json. Auth via PUSHOVER_APP_TOKEN and PUSHOVER_USER_KEY env vars.",
        ],
        [
            "PreRollTracker",
            "GitHub Releases",
            "REST API (GitHub API v3)",
            "Outbound: uploads AES-256 encrypted database backups as release assets for off-site disaster recovery",
            "https://api.github.com/repos/{owner}/{repo}/releases. Auth via GITHUB_TOKEN env var. Repo set via GITHUB_REPO.",
        ],
        [
            "ApexAPI",
            "Windows Printer",
            "Platform API (win32print / lpr)",
            "Outbound: sends generated PDF order forms and labels to configured local or network printer",
            "win32print (Windows) or lpr (macOS/Linux). Printer selected via selected_printer config key in apex_config.json.",
        ],
    ]

    add_rows(ws, headers, data)
    style_sheet(ws, len(headers))


def build_config_options_sheet(ws):
    headers = ["App", "Key", "Type", "Default", "Description", "Where Set"]

    data = [
        # ── PreRollTracker: Environment Variables ──
        ["PreRollTracker", "FLASK_ENV", "str", "production", "Flask environment mode: development or production. Controls debug mode, reloader, and logging verbosity.", "Environment variable (.env file or system)"],
        ["PreRollTracker", "SECRET_KEY", "str", "(generated)", "Flask session secret key for signing cookies and CSRF tokens. Auto-generated if not set.", "Environment variable"],
        ["PreRollTracker", "ADMIN_PASSWORD_HASH", "str", "(none)", "Bcrypt hash of the admin password. Used for admin login authentication.", "Environment variable"],
        ["PreRollTracker", "PUSHOVER_APP_TOKEN", "str", "(none)", "Pushover application API token for sending push notifications on failures.", "Environment variable"],
        ["PreRollTracker", "PUSHOVER_USER_KEY", "str", "(none)", "Pushover user/group key identifying notification recipients.", "Environment variable"],
        ["PreRollTracker", "BACKUP_ENCRYPTION_KEY", "str", "(none)", "Fernet-compatible key for AES-256 encrypting database backups.", "Environment variable"],
        ["PreRollTracker", "GITHUB_TOKEN", "str", "(none)", "GitHub personal access token for uploading backup releases.", "Environment variable"],
        ["PreRollTracker", "GITHUB_REPO", "str", "(none)", "GitHub repository in owner/repo format for backup uploads.", "Environment variable"],
        ["PreRollTracker", "DATABASE_URL", "str", "sqlite:///preroll.db", "Database connection URI. Defaults to local SQLite file.", "Environment variable"],
        ["PreRollTracker", "LOG_LEVEL", "str", "INFO", "Python logging level: DEBUG, INFO, WARNING, ERROR, CRITICAL.", "Environment variable"],
        ["PreRollTracker", "HOST", "str", "0.0.0.0", "Network interface to bind the Flask server to.", "Environment variable"],
        ["PreRollTracker", "PORT", "int", "5000", "TCP port for the Flask server.", "Environment variable"],
        ["PreRollTracker", "SENTRY_DSN", "str", "(none)", "Sentry Data Source Name for error tracking integration.", "Environment variable"],
        ["PreRollTracker", "SENTRY_ENV", "str", "production", "Sentry environment tag for distinguishing dev/staging/prod errors.", "Environment variable"],

        # ── PreRollTracker: Settings Table ──
        ["PreRollTracker", "backup_enabled", "bool", "true", "Enable/disable automated backup scheduling.", "Settings table (Admin UI)"],
        ["PreRollTracker", "backup_interval_hours", "int", "6", "Hours between automated backups.", "Settings table (Admin UI)"],
        ["PreRollTracker", "backup_retention_days", "int", "30", "Days to retain local backup files before cleanup.", "Settings table (Admin UI)"],
        ["PreRollTracker", "low_stock_threshold", "int", "100", "Default low-stock alert threshold for inventory items.", "Settings table (Admin UI)"],
        ["PreRollTracker", "achievement_notifications", "bool", "true", "Enable/disable achievement notification popups for workers.", "Settings table (Admin UI)"],
        ["PreRollTracker", "learning_enabled", "bool", "true", "Enable/disable the production rate learning processor.", "Settings table (Admin UI)"],
        ["PreRollTracker", "default_target_quantity", "int", "500", "Default target quantity when creating new batches.", "Settings table (Admin UI)"],
        ["PreRollTracker", "qc_required", "bool", "false", "Require QC check before a batch can be marked complete.", "Settings table (Admin UI)"],
        ["PreRollTracker", "wholesale_hold_expiry_hours", "int", "48", "Hours before an unconfirmed wholesale hold auto-expires.", "Settings table (Admin UI)"],

        # ── ApexAPI: Config File (apex_config.json) ──
        ["ApexAPI", "api_token", "str", "(none)", "Apex Trading API bearer token for authentication.", "apex_config.json"],
        ["ApexAPI", "dashboard_api_key", "str", "(none)", "API key for authenticating with PreRollTracker dashboard endpoints.", "apex_config.json"],
        ["ApexAPI", "dashboard_url", "str", "https://himomstats.online", "Base URL of the PreRollTracker dashboard API.", "apex_config.json"],
        ["ApexAPI", "license_number", "str", "(none)", "Cannabis license number used for METRC and regulatory tagging.", "apex_config.json"],
        ["ApexAPI", "selected_printer", "str", "(none)", "Name of the printer to use for order form printing.", "apex_config.json / GUI"],
        ["ApexAPI", "auto_refresh_enabled", "bool", "true", "Enable auto-refresh of the order list.", "apex_config.json / GUI toggle"],
        ["ApexAPI", "auto_refresh_interval", "int", "5", "Seconds between auto-refresh cycles.", "apex_config.json / GUI"],
        ["ApexAPI", "excluded_statuses", "list[str]", '["cancelled", "void"]', "Order statuses to hide from the main order list.", "apex_config.json"],
        ["ApexAPI", "cache_warming.enabled", "bool", "true", "Enable proactive cache warming.", "apex_config.json"],
        ["ApexAPI", "cache_warming.strategies", "dict", "(per-type intervals)", "Map of data types to warming intervals and priorities.", "apex_config.json"],
        ["ApexAPI", "cache_warming.max_concurrent", "int", "3", "Maximum concurrent cache warming tasks.", "apex_config.json"],
        ["ApexAPI", "batch_inventory_sync.enabled", "bool", "true", "Enable automatic batch inventory sync to Apex API.", "apex_config.json"],
        ["ApexAPI", "batch_inventory_sync.interval_seconds", "int", "5", "Seconds between inventory sync cycles.", "apex_config.json"],
        ["ApexAPI", "batch_inventory_sync.retry_count", "int", "3", "Number of retries on sync failure before giving up.", "apex_config.json"],
        ["ApexAPI", "pre_roll_packing_list.google_sheet_id", "str", "(none)", "Google Sheet ID for pre-roll packing list sync.", "apex_config.json"],
        ["ApexAPI", "pre_roll_packing_list.sync_interval_minutes", "int", "15", "Minutes between Google Sheets sync cycles.", "apex_config.json"],
        ["ApexAPI", "pre_roll_packing_list.default_view", "str", "by_store", "Default packing list view: by_store or by_product.", "apex_config.json / GUI"],
        ["ApexAPI", "edibles.track_sell_through", "bool", "true", "Enable sell-through rate tracking for edibles.", "apex_config.json"],
        ["ApexAPI", "edibles.reorder_threshold", "float", "0.3", "Sell-through rate below which a reorder is recommended.", "apex_config.json"],
        ["ApexAPI", "store_aliases", "dict", "{}", "Map of canonical store names to display aliases.", "apex_config.json / Store Management UI"],
        ["ApexAPI", "delivery_exclusions", "list[str]", "[]", "Store names excluded from delivery routing.", "apex_config.json / Store Management UI"],
        ["ApexAPI", "order_form.template_path", "str", "templates/order_form.html", "Path to the PDF order form Jinja2 template.", "apex_config.json"],
        ["ApexAPI", "order_form.output_dir", "str", "output/forms", "Directory for generated PDF order forms.", "apex_config.json"],
        ["ApexAPI", "vape_cart.categories", "list[str]", '["510", "Pod", "AIO", "Disposable"]', "Recognized vape cart form factor categories.", "apex_config.json"],
    ]

    add_rows(ws, headers, data)
    style_sheet(ws, len(headers))


def main():
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Build each sheet
    sheet_builders = [
        ("Pages & Screens", build_pages_sheet),
        ("API Endpoints", build_api_endpoints_sheet),
        ("Database Tables", build_database_tables_sheet),
        ("Background Jobs", build_background_jobs_sheet),
        ("Integrations", build_integrations_sheet),
        ("Config Options", build_config_options_sheet),
    ]

    for sheet_name, builder in sheet_builders:
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_properties.tabColor = TAB_COLORS.get(sheet_name, "000000")
        builder(ws)

    output_path = "/Users/chrisgillis/PycharmProjects/HiMoM/docs/inventory.xlsx"
    wb.save(output_path)
    print(f"Workbook saved to {output_path}")

    # Print summary
    for ws in wb.worksheets:
        print(f"  Sheet '{ws.title}': {ws.max_row - 1} data rows, {ws.max_column} columns")


if __name__ == "__main__":
    main()

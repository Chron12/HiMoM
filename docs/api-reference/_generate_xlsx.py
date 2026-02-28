#!/usr/bin/env python3
"""
Generate api-reference.xlsx and data-models.xlsx for PreRollTracker API documentation.
"""

import sys
sys.path.insert(0, "/Users/chrisgillis/PycharmProjects/HiMoM/.venv/lib/python3.14/site-packages")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def style_header_row(ws, num_cols):
    """Apply consistent header styling."""
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A5676", end_color="1A5676", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def style_data_rows(ws, num_rows, num_cols):
    """Apply alternating row colors and borders."""
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alt_fill = PatternFill(start_color="F0F7FA", end_color="F0F7FA", fill_type="solid")
    for row in range(2, num_rows + 2):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row % 2 == 0:
                cell.fill = alt_fill


def auto_width(ws, num_cols, max_width=60):
    """Auto-fit column widths based on content."""
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    for line in lines:
                        max_len = max(max_len, len(line))
        adjusted = min(max_len + 4, max_width)
        ws.column_dimensions[get_column_letter(col)].width = max(adjusted, 12)


# =============================================================================
# 1. API Reference Spreadsheet
# =============================================================================

def generate_api_reference():
    wb = Workbook()
    ws = wb.active
    ws.title = "API Endpoints"

    headers = ["Method", "Path", "Description", "Auth", "Parameters", "Response Type", "Rate Limit"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    endpoints = [
        # AUTH
        ("GET", "/login", "Render admin login page", "None", "None", "HTML", "None"),
        ("POST", "/login", "Authenticate and establish session", "None", "password (form), remember_me (form), csrf_token (form)", "302 Redirect", "5/min"),
        ("GET", "/logout", "Log out and clear session", "None", "None", "302 Redirect", "None"),
        ("GET", "/forgot-password", "Render password recovery page", "None", "None", "HTML", "None"),
        ("POST", "/forgot-password", "Verify recovery key", "None", "recovery_key (form), csrf_token (form)", "302 Redirect", "2/min"),
        ("GET", "/reset-password", "Render password reset form", "None", "token (query)", "HTML", "None"),
        ("POST", "/reset-password", "Set new password after recovery", "None", "token (form), new_password (form), confirm_password (form), csrf_token (form)", "HTML", "5/min"),

        # BATCHES
        ("GET", "/api/data", "Get all active batches", "API Key or Session", "include_history (query, bool, default: false)", "JSON Array", "None"),
        ("GET", "/api/archive", "Get all archived batches", "API Key or Session", "None", "JSON Array", "None"),
        ("GET", "/api/batches/last-updated", "Polling: latest batch change timestamp", "API Key or Session", "None", "JSON {timestamp}", "None"),
        ("GET", "/api/batch/<batch_id>/rate-history", "Rate history for a batch", "API Key or Session", "batch_id (URL path)", "JSON Object", "None"),
        ("POST", "/api/batch/<batch_id>/counts", "Save production counts", "API Key or Session", "batch_id (URL path); Body: counts_0_5 (int), counts_0_7 (int), counts_1_0 (int)", "JSON {ok: true}", "None"),
        ("POST", "/api/batch/<batch_id>/plan", "Save planned counts", "API Key or Session", "batch_id (URL path); Body: planned_0_5 (int), planned_0_7 (int), planned_1_0 (int), plan_use_grams (float)", "JSON {ok, adjusted, planned_*}", "None"),
        ("GET", "/api/strain/<strain_name>/rate-projection", "Rate projection for strain", "API Key or Session", "strain_name (URL path)", "JSON Object", "None"),
        ("POST", "/api/archive/<batch_id>", "Toggle batch archive status", "API Key or Session", "batch_id (URL path); Body: archive (bool, default: true)", "JSON {status, action}", "None"),
        ("POST", "/api/public/archive/<batch_id>", "Archive completed batch (stage 7 only)", "API Key or Session", "batch_id (URL path)", "JSON {status, action}", "None"),
        ("GET", "/api/allocation-preview", "Preview allocation for weight", "Session Only", "weight (query, float, default: 1000)", "JSON Object", "None"),
        ("POST", "/api/reorder", "Reorder batches", "API Key or Session", "Body: batch_ids (array of strings)", "JSON {status}", "None"),
        ("POST", "/api/batch/<batch_id>/centrifuge", "Save centrifuge settings for batch", "Session Only", "batch_id (URL path); Body: centrifuge_rpm (int), centrifuge_time_seconds (int), centrifuge_cycles (int), centrifuge_fill_gauge_cycle1 (int), centrifuge_fill_gauge_cycle2 (int)", "JSON {status}", "None"),
        ("POST", "/api/update-progress/<batch_id>", "Update production progress", "API Key or Session", "batch_id (URL path); Body: counts_0_5 (int), counts_0_7 (int), counts_1_0 (int)", "JSON {status, data}", "None"),

        # INVENTORY
        ("GET", "/api/inventory/", "Get current inventory data", "API Key or Session", "None", "JSON Object", "None"),
        ("POST", "/api/inventory/update", "Update inventory counts", "API Key or Session", "Body: size (string: 0_5/0_7/1_0), box_count (int), individual_papers (int)", "JSON {status, inventory}", "None"),
        ("POST", "/api/inventory/settings", "Update inventory settings", "API Key or Session", "Body: thresholds (object), papers_per_box (object)", "JSON {status}", "None"),

        # FINISHED GOODS
        ("GET", "/api/finished-goods/", "List finished goods packages", "API Key or Session", "include_archived (query, bool), strain (query, string), status (query, string), search (query, string)", "JSON {packages, total, summary, settings}", "None"),
        ("GET", "/api/finished-goods/last-updated", "Polling: latest FG change timestamp", "API Key or Session", "None", "JSON {timestamp}", "None"),
        ("GET", "/api/finished-goods/summary", "Summary statistics", "API Key or Session", "None", "JSON Object", "None"),
        ("GET", "/api/finished-goods/history", "Recent history across all packages", "API Key or Session", "limit (query, int, default: 100)", "JSON {history, total}", "None"),
        ("GET", "/api/finished-goods/calculator", "Pre-roll estimates for grams", "API Key or Session", "grams (query, float, required)", "JSON Object", "None"),
        ("GET", "/api/finished-goods/<metrc_number>", "Get specific package", "API Key or Session", "metrc_number (URL path)", "JSON Object", "None"),
        ("POST", "/api/finished-goods/", "Add new package", "API Key or Session", "Body: metrc_number (string, req), strain (string, req), grams (float, req), notes (string), source_batch_id (string)", "JSON {success, package} (201)", "None"),
        ("PUT", "/api/finished-goods/<metrc_number>", "Update package fields", "API Key or Session", "metrc_number (URL path); Body: strain (string), notes (string), source_batch_id (string), reason (string)", "JSON {success, updated_fields, package}", "None"),
        ("DELETE", "/api/finished-goods/<metrc_number>", "Delete package", "API Key or Session", "metrc_number (URL path)", "JSON", "None"),
        ("POST", "/api/finished-goods/<metrc>/deduct", "Deduct inventory from package", "API Key or Session", "metrc_number (URL); Body: grams (float) OR units (int) + unit_size (string); reason (string)", "JSON {success, ...}", "None"),
        ("POST", "/api/finished-goods/<metrc>/add", "Add inventory to package", "API Key or Session", "metrc_number (URL); Body: grams (float) OR units (int) + unit_size (string); reason (string)", "JSON {success, ...}", "None"),
        ("POST", "/api/finished-goods/<metrc>/physical-override", "Set physical inventory override", "API Key or Session", "metrc_number (URL); Body: physical_grams (float|null), reason (string)", "JSON {success, effective_grams, ...}", "None"),

        # WHOLESALE
        ("GET", "/api/wholesale/inventory", "Get wholesale inventory", "Session Only", "None", "JSON {success, strains}", "None"),
        ("POST", "/api/wholesale/hold", "Create inventory hold", "Session Only", "Body: metrc_number (string, req), sku_name (string, req), quantity (int, req), notes (string)", "JSON {success, hold_id}", "None"),
        ("DELETE", "/api/wholesale/hold/<hold_id>", "Release hold", "Session Only", "hold_id (URL path)", "JSON {success}", "None"),
        ("GET", "/api/wholesale/holds", "List all active holds", "Session Only", "metrc_number (query, string, optional)", "JSON {success, holds}", "None"),
        ("GET", "/api/wholesale/last-updated", "Polling: wholesale change fingerprint", "Session Only", "None", "JSON {fingerprint}", "None"),

        # CENTRIFUGE
        ("POST", "/api/centrifuge/calculate", "Force calculation", "API Key or Session", "Body: rpm (int, req), centrifuge_type (string, default: silver_bullet)", "JSON Object", "None"),
        ("POST", "/api/centrifuge/compare", "Machine comparison", "API Key or Session", "Body: rpm (int, req), source_centrifuge (string, default: silver_bullet)", "JSON Object", "None"),
        ("GET", "/api/centrifuge/curve-data", "Force curve data for graphing", "API Key or Session", "centrifuge (query, string), rpm_min (query, int), rpm_max (query, int), step (query, int)", "JSON {success, data}", "None"),
        ("GET", "/api/centrifuge/settings-guide", "Recommended settings", "API Key or Session", "material (query, string, default: standard)", "JSON {success, recommendations, safety_zones, centrifuges}", "None"),
        ("POST", "/api/centrifuge/impulse-calculate", "Impulse calculation", "API Key or Session", "Body: rpm (int, req), time_seconds (float), weight_grams (float), centrifuge_type (string)", "JSON {success, data}", "None"),
        ("POST", "/api/centrifuge/target-match", "Target impulse matching", "API Key or Session", "Body: target_impulse (float), adjust_mode (string: rpm|time), rpm (int), time_seconds (float), weight_grams (float), centrifuge_type (string)", "JSON {success, data}", "None"),
        ("POST", "/api/centrifuge/batch-comparison", "Batch weight comparison", "API Key or Session", "Body: rpm (int), time_seconds (float), centrifuge_type (string), batch_weights (array of floats)", "JSON {success, data}", "None"),
        ("GET", "/api/centrifuge/impulse-zones", "Get impulse zone definitions", "API Key or Session", "None", "JSON {success, data}", "None"),

        # SNAPSHOTS
        ("GET", "/api/snapshots/history", "Available snapshots", "Session Only", "None", "JSON {snapshots}", "None"),
        ("GET", "/api/snapshots/<timestamp>", "Specific snapshot", "Session Only", "timestamp (URL path)", "JSON Object", "None"),
        ("POST", "/api/snapshots/compare-detailed", "Compare two snapshots", "Session Only", "Body: timestamp1 (string, req), timestamp2 (string, req), store_filter (string)", "JSON {comparisons}", "None"),
        ("POST", "/api/snapshots/insights", "Generate insights", "Session Only", "Body: timestamp1 (string, req), timestamp2 (string, req)", "JSON {insights}", "None"),
        ("POST", "/api/snapshots/export-csv", "Export snapshot to CSV", "Session Only", "Body: timestamp (string, req)", "JSON {csv, filename}", "None"),

        # MISC
        ("GET", "/api/version", "App version (public)", "None", "None", "JSON {version}", "None"),
        ("GET", "/api/backup-status", "Backup health", "API Key or Session", "None", "JSON Object", "None"),
        ("GET", "/api/settings", "Current settings", "API Key or Session", "None", "JSON {allocation, tare_weights, test_alert_hours, work_schedule}", "None"),
        ("GET", "/api/api-key", "View API key", "Session Only", "None", "JSON {api_key, usage}", "None"),
        ("POST", "/api/api-key/regenerate", "Generate new API key", "Session Only", "None", "JSON {api_key, message}", "None"),
        ("GET", "/api/overview", "Production overview", "API Key or Session", "None", "JSON {stage_counts, production_stats, priority_batches, total_active}", "None"),
        ("GET", "/api/audit", "Audit trail data", "API Key or Session", "None", "JSON Array", "None"),
        ("GET", "/api/production-history", "Historical production data", "API Key or Session", "days (query, int, default: 7, range: 1-90)", "JSON Object", "None"),
        ("POST", "/api/dismiss-inventory-alert/<size>", "Dismiss single alert", "Session Only", "size (URL path: 0_5/0_7/1_0)", "JSON {status}", "None"),
        ("POST", "/api/dismiss-inventory-alerts", "Dismiss all alerts", "Session Only", "None", "JSON {status}", "None"),
        ("GET", "/api/apex-sync-status", "ApexAPI sync status", "API Key or Session", "None", "JSON {last_sync_requested}", "None"),
        ("POST", "/api/apex-sync-trigger", "Manual sync trigger", "API Key or Session", "None", "JSON {success, last_sync_requested}", "None"),
        ("GET", "/api/pushover/settings", "Pushover config", "Session Only", "None", "JSON Object", "None"),
        ("PUT", "/api/pushover/settings", "Update Pushover config", "Session Only", "Body: enabled (bool), user_key (string), warning_grams (float), critical_grams (float), cooldown_hours (int)", "JSON {success, settings}", "None"),
        ("POST", "/api/pushover/test", "Test notification", "Session Only", "level (query, string, default: warning)", "JSON {success, message}", "None"),
        ("POST", "/api/pushover/check-alerts", "Check and send stock alerts", "Session Only", "None", "JSON {success, result}", "None"),
        ("GET", "/api/centrifuge-recommendations-unified/<strain>", "Unified centrifuge recommendations", "API Key or Session", "strain_name (URL); grind_size (query), priority_mode (query), temporal (query), target_batch (query), days_since_harvest (query)", "JSON Object", "None"),
        ("GET", "/api/centrifuge-trends/<strain_name>", "Centrifuge trends for strain", "API Key or Session", "strain_name (URL); grind_size (query), target_batch (query), days_since_harvest (query)", "JSON Object", "None"),
        ("GET", "/api/centrifuge-history/<strain_name>", "Centrifuge setting change history", "API Key or Session", "strain_name (URL path)", "JSON {strain, history, total_changes}", "None"),
        ("POST", "/api/weight-check/<batch_id>", "Log weight measurement", "Session Only", "batch_id (URL); Body: size (string: 0.5g/0.7g/1.0g), sample_weight (float), count_at_measurement (int)", "JSON {success, analysis}", "None"),
        ("GET", "/api/weight-analytics/<batch_id>", "Weight tracking analytics", "Session Only", "batch_id (URL path)", "JSON Object", "None"),
        ("GET", "/api/recommendations", "Latest inventory recommendations", "Session Only", "None", "JSON Object", "None"),
        ("POST", "/api/recommendations/upload", "Upload CSV for recommendations", "Session Only", "files[] (multipart/form-data, CSV files)", "JSON {success, strain_count, store_count, timestamp}", "None"),

        # PWA
        ("GET", "/manifest.json", "PWA manifest", "None", "None", "JSON", "None"),
        ("GET", "/sw.js", "Service worker", "None", "None", "JavaScript", "None"),
        ("GET", "/icon-192.png", "192x192 app icon (SVG)", "None", "None", "SVG Image", "None"),
        ("GET", "/icon-512.png", "512x512 app icon (SVG)", "None", "None", "SVG Image", "None"),
        ("GET", "/favicon.ico", "Favicon (SVG)", "None", "None", "SVG Image", "None"),
    ]

    for row_idx, ep in enumerate(endpoints, 2):
        for col_idx, value in enumerate(ep, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    style_header_row(ws, len(headers))
    style_data_rows(ws, len(endpoints), len(headers))
    auto_width(ws, len(headers))

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(endpoints) + 1}"

    output_path = "/Users/chrisgillis/PycharmProjects/HiMoM/docs/api-reference/api-reference.xlsx"
    wb.save(output_path)
    print(f"Generated: {output_path}")


# =============================================================================
# 2. Data Models Spreadsheet
# =============================================================================

def generate_data_models():
    wb = Workbook()

    # -------------------------------------------------------------------------
    # Sheet 1: PreRollTracker Models
    # -------------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "PreRollTracker Models"

    headers = ["Model", "Field", "Type", "Default", "Constraints", "Description"]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    models = [
        # Batches
        ("Batch", "id", "TEXT", "UUID", "PRIMARY KEY", "Unique batch identifier"),
        ("Batch", "strain", "TEXT", "", "NOT NULL", "Cannabis strain name"),
        ("Batch", "input_grams", "REAL", "", "NOT NULL", "Total input weight in grams"),
        ("Batch", "target_grams_each", "REAL", "", "NOT NULL", "Target weight per pre-roll"),
        ("Batch", "produced", "INTEGER", "0", "", "Total units produced"),
        ("Batch", "grams_used", "REAL", "0.0", "", "Grams consumed in production"),
        ("Batch", "grams_ground", "REAL", "0.0", "", "Grams ground so far"),
        ("Batch", "stage", "INTEGER", "0", "0-7", "Production stage index"),
        ("Batch", "counts_0_5", "INTEGER", "0", "", "0.5g pre-rolls produced"),
        ("Batch", "counts_0_7", "INTEGER", "0", "", "0.7g pre-rolls produced"),
        ("Batch", "counts_1_0", "INTEGER", "0", "", "1.0g pre-rolls produced"),
        ("Batch", "_last_inventory_consumed_0_5", "INTEGER", "0", "", "Internal: last consumed 0.5g papers"),
        ("Batch", "_last_inventory_consumed_0_7", "INTEGER", "0", "", "Internal: last consumed 0.7g papers"),
        ("Batch", "_last_inventory_consumed_1_0", "INTEGER", "0", "", "Internal: last consumed 1.0g papers"),
        ("Batch", "planned_0_5", "INTEGER", "0", "", "Planned 0.5g count"),
        ("Batch", "planned_0_7", "INTEGER", "0", "", "Planned 0.7g count"),
        ("Batch", "planned_1_0", "INTEGER", "0", "", "Planned 1.0g count"),
        ("Batch", "plan_use_grams", "REAL", "0.0", "", "Grams allocated for plan"),
        ("Batch", "labels_printed", "INTEGER", "0", "", "Number of labels printed"),
        ("Batch", "display_order", "INTEGER", "0", "", "Sort order for UI display"),
        ("Batch", "production_start_time", "TEXT", "NULL", "ISO 8601", "When production began"),
        ("Batch", "production_end_time", "TEXT", "NULL", "ISO 8601", "When production ended"),
        ("Batch", "production_duration_hours", "REAL", "NULL", "", "Total production hours"),
        ("Batch", "archived", "INTEGER", "0", "0 or 1", "Whether batch is archived"),
        ("Batch", "harvest_date", "TEXT", "NULL", "ISO 8601", "Harvest date of source material"),
        ("Batch", "bay_number", "TEXT", "NULL", "", "Bay/location number"),
        ("Batch", "thc_percent", "REAL", "NULL", "0-100", "THC percentage from lab test"),
        ("Batch", "thca_percent", "REAL", "NULL", "0-100", "THCA percentage from lab test"),
        ("Batch", "tac_percent", "REAL", "NULL", "0-100", "Total Active Cannabinoids percentage"),
        ("Batch", "cbd_percent", "REAL", "NULL", "0-100", "CBD percentage from lab test"),
        ("Batch", "origin_metrc_number", "TEXT", "NULL", "", "Source METRC package number"),
        ("Batch", "new_metrc_created", "INTEGER", "0", "0 or 1", "Whether new METRC pkg created"),
        ("Batch", "centrifuge_rpm", "INTEGER", "NULL", "500-2500", "Centrifuge RPM setting"),
        ("Batch", "centrifuge_time_seconds", "INTEGER", "NULL", "1-90", "Centrifuge time in seconds"),
        ("Batch", "centrifuge_cycles", "INTEGER", "NULL", "1-3", "Number of centrifuge cycles"),
        ("Batch", "centrifuge_fill_gauge_cycle1", "REAL", "NULL", "1-12", "Fill gauge reading cycle 1"),
        ("Batch", "centrifuge_fill_gauge_cycle2", "REAL", "NULL", "1-12", "Fill gauge reading cycle 2"),
        ("Batch", "centrifuge_machine", "TEXT", "NULL", "", "Centrifuge machine name"),
        ("Batch", "centrifuge_settings_by_size", "TEXT (JSON)", "NULL", "", "Per-size centrifuge settings"),
        ("Batch", "new_metrc_number", "TEXT", "NULL", "", "Newly created METRC number"),
        ("Batch", "metrc_transferred", "INTEGER", "0", "0 or 1", "METRC transfer completed"),
        ("Batch", "metrc_batch_created", "INTEGER", "0", "0 or 1", "METRC batch created"),
        ("Batch", "testing_status", "TEXT", "none", "none/sample_created/sample_sent/results_received/test_passed/test_failed", "Current testing status"),
        ("Batch", "test_package_id", "TEXT", "NULL", "", "Test sample package ID"),
        ("Batch", "testing_lab", "TEXT", "NULL", "", "Testing laboratory name"),
        ("Batch", "test_date", "TEXT", "NULL", "ISO 8601", "Test date"),
        ("Batch", "test_results", "TEXT (JSON)", "NULL", "", "Serialized test results"),
        ("Batch", "testing_notes", "TEXT", "NULL", "", "Testing notes"),
        ("Batch", "packaging_status", "TEXT", "available", "available/waiting_materials/waiting_containers/blocked_other", "Packaging dependency status"),
        ("Batch", "packaging_blocker_type", "TEXT", "NULL", "", "Type of packaging blocker"),
        ("Batch", "packaging_notes", "TEXT", "NULL", "", "Packaging notes"),
        ("Batch", "packaging_blocked_since", "TEXT", "NULL", "ISO 8601", "When packaging was blocked"),
        ("Batch", "cached_production_rate", "REAL", "NULL", "", "Cached units/hour rate"),
        ("Batch", "rate_calculation_time", "TEXT", "NULL", "ISO 8601", "When rate was last calculated"),
        ("Batch", "rate_history", "TEXT (JSON)", "NULL", "", "Array of historical rate data points"),
        ("Batch", "rate_metrics", "TEXT (JSON)", "NULL", "", "Rate analysis metrics"),
        ("Batch", "last_rate_update", "TEXT", "NULL", "ISO 8601", "Last rate update timestamp"),
        ("Batch", "production_sessions", "TEXT (JSON)", "NULL", "", "Work session tracking data"),
        ("Batch", "current_session_start", "TEXT", "NULL", "ISO 8601", "Current work session start"),
        ("Batch", "total_work_hours", "REAL", "0.0", "", "Total accumulated work hours"),
        ("Batch", "paper_size_override", "TEXT", "NULL", "", "Override paper size configuration"),
        ("Batch", "is_infused", "INTEGER", "0", "0 or 1", "Whether this is an infused product"),
        ("Batch", "source_materials", "TEXT (JSON)", "NULL", "", "Source materials for infused products"),
        ("Batch", "weight_log_0_5", "TEXT (JSON)", "NULL", "", "Weight measurements for 0.5g"),
        ("Batch", "weight_log_0_7", "TEXT (JSON)", "NULL", "", "Weight measurements for 0.7g"),
        ("Batch", "weight_log_1_0", "TEXT (JSON)", "NULL", "", "Weight measurements for 1.0g"),
        ("Batch", "running_yield_0_5", "REAL", "0.0", "", "Running yield for 0.5g"),
        ("Batch", "running_yield_0_7", "REAL", "0.0", "", "Running yield for 0.7g"),
        ("Batch", "running_yield_1_0", "REAL", "0.0", "", "Running yield for 1.0g"),
        ("Batch", "final_weight_0_5", "REAL", "0.0", "", "Final batch weight for 0.5g"),
        ("Batch", "final_weight_0_7", "REAL", "0.0", "", "Final batch weight for 0.7g"),
        ("Batch", "final_weight_1_0", "REAL", "0.0", "", "Final batch weight for 1.0g"),
        ("Batch", "grind_size", "TEXT", "NULL", "fine/medium/coarse", "Grind size used"),
        ("Batch", "packaged_singles", "INTEGER", "0", "", "Singles packaged"),
        ("Batch", "packaged_6_packs", "INTEGER", "0", "", "6-packs packaged"),
        ("Batch", "packaged_12_packs", "INTEGER", "0", "", "12-packs packaged"),
        ("Batch", "loose_tupperware", "INTEGER", "0", "", "Units in loose tupperware"),

        # Audit Log
        ("AuditLog", "id", "INTEGER", "AUTO", "PRIMARY KEY AUTOINCREMENT", "Unique entry ID"),
        ("AuditLog", "batch", "TEXT", "", "NOT NULL", "Batch ID reference"),
        ("AuditLog", "strain", "TEXT", "", "NOT NULL", "Strain name at time of change"),
        ("AuditLog", "ts", "TEXT", "", "NOT NULL, ISO 8601", "Timestamp of the change"),
        ("AuditLog", "field", "TEXT", "", "NOT NULL", "Name of the changed field"),
        ("AuditLog", "old", "TEXT", "NULL", "", "Previous value (serialized)"),
        ("AuditLog", "new", "TEXT", "NULL", "", "New value (serialized)"),

        # Finished Goods
        ("FinishedGoods", "metrc_number", "TEXT", "", "PRIMARY KEY", "METRC package number"),
        ("FinishedGoods", "strain", "TEXT", "", "NOT NULL", "Cannabis strain name"),
        ("FinishedGoods", "initial_grams", "REAL", "", "NOT NULL", "Initial weight when created"),
        ("FinishedGoods", "current_grams", "REAL", "", "NOT NULL", "Current remaining weight"),
        ("FinishedGoods", "status", "TEXT", "active", "active/depleted/archived", "Package status"),
        ("FinishedGoods", "created_date", "TEXT", "NULL", "ISO 8601", "When package was created"),
        ("FinishedGoods", "updated_date", "TEXT", "NULL", "ISO 8601", "Last modification time"),
        ("FinishedGoods", "notes", "TEXT", "NULL", "", "User notes"),
        ("FinishedGoods", "source_batch_id", "TEXT", "NULL", "FK -> batches(id)", "Source production batch"),
        ("FinishedGoods", "grams_ordered", "REAL", "0.0", "", "Grams ordered from package"),
        ("FinishedGoods", "grams_packed", "REAL", "0.0", "", "Grams packed from package"),
        ("FinishedGoods", "grams_packed_lifetime", "REAL", "0.0", "", "Lifetime total grams packed"),
        ("FinishedGoods", "grams_fulfilled", "REAL", "0.0", "", "Grams fulfilled/delivered"),
        ("FinishedGoods", "sku_breakdown", "TEXT (JSON)", "NULL", "", "SKU breakdown data"),
        ("FinishedGoods", "apex_auto_inventory", "INTEGER", "0", "0 or 1", "Auto-inventory via Apex"),
        ("FinishedGoods", "apex_units", "TEXT (JSON)", "NULL", "", "Calculated Apex unit counts"),
        ("FinishedGoods", "apex_sku_settings", "TEXT (JSON)", "NULL", "", "Apex SKU configurations"),
        ("FinishedGoods", "processed_decrements", "TEXT (JSON)", "NULL", "", "Tracked decrement operations"),
        ("FinishedGoods", "custom_skus", "TEXT (JSON)", "NULL", "", "Custom SKU definitions"),
        ("FinishedGoods", "physical_grams_override", "REAL", "NULL", "", "Physical count override"),

        # Finished Goods History
        ("FinishedGoodsHistory", "id", "INTEGER", "AUTO", "PRIMARY KEY AUTOINCREMENT", "Unique entry ID"),
        ("FinishedGoodsHistory", "metrc_number", "TEXT", "", "NOT NULL", "Package METRC number"),
        ("FinishedGoodsHistory", "timestamp", "TEXT", "", "NOT NULL, ISO 8601", "When the change occurred"),
        ("FinishedGoodsHistory", "change_type", "TEXT", "", "NOT NULL", "Type: created/deduct/add/update/archive"),
        ("FinishedGoodsHistory", "current_grams", "REAL", "", "NOT NULL", "Grams after change"),
        ("FinishedGoodsHistory", "grams_ordered", "REAL", "0.0", "", "Ordered grams at time"),
        ("FinishedGoodsHistory", "grams_packed", "REAL", "0.0", "", "Packed grams at time"),
        ("FinishedGoodsHistory", "grams_fulfilled", "REAL", "0.0", "", "Fulfilled grams at time"),
        ("FinishedGoodsHistory", "status", "TEXT", "NULL", "", "Package status at time"),
        ("FinishedGoodsHistory", "details", "TEXT (JSON)", "NULL", "", "Additional change details"),

        # Inventory
        ("Inventory", "size", "TEXT", "", "PRIMARY KEY", "Size key: 0_5, 0_7, 1_0"),
        ("Inventory", "boxes", "INTEGER", "0", "", "Number of full boxes"),
        ("Inventory", "individual_papers", "INTEGER", "0", "", "Loose individual papers"),
        ("Inventory", "low_threshold", "INTEGER", "5", "", "Low-stock alert threshold (boxes)"),
        ("Inventory", "last_updated", "TEXT", "NULL", "ISO 8601", "Last update timestamp"),

        # Inventory Usage
        ("InventoryUsage", "id", "INTEGER", "AUTO", "PRIMARY KEY AUTOINCREMENT", "Unique entry ID"),
        ("InventoryUsage", "size", "TEXT", "", "NOT NULL", "Size key: 0_5, 0_7, 1_0"),
        ("InventoryUsage", "timestamp", "TEXT", "", "NOT NULL, ISO 8601", "When papers were consumed"),
        ("InventoryUsage", "papers_used", "INTEGER", "", "NOT NULL", "Number of papers consumed"),
        ("InventoryUsage", "batch_id", "TEXT", "NULL", "", "Source batch ID"),

        # Settings
        ("Settings", "key", "TEXT", "", "PRIMARY KEY", "Setting name"),
        ("Settings", "value", "TEXT", "", "NOT NULL", "Setting value (JSON serialized)"),
        ("Settings", "updated_at", "TEXT", "", "NOT NULL, ISO 8601", "Last update timestamp"),

        # Wholesale Holds
        ("WholesaleHold", "id", "TEXT", "UUID", "PRIMARY KEY", "Unique hold identifier"),
        ("WholesaleHold", "metrc_number", "TEXT", "", "NOT NULL, FK -> finished_goods ON DELETE CASCADE", "Package METRC number"),
        ("WholesaleHold", "sku_name", "TEXT", "", "NOT NULL", "SKU being held"),
        ("WholesaleHold", "quantity", "INTEGER", "", "NOT NULL", "Units held"),
        ("WholesaleHold", "created_date", "TEXT", "", "NOT NULL, ISO 8601", "When hold was created"),
        ("WholesaleHold", "notes", "TEXT", "NULL", "", "Hold notes"),
    ]

    for row_idx, m in enumerate(models, 2):
        for col_idx, value in enumerate(m, 1):
            ws1.cell(row=row_idx, column=col_idx, value=value)

    style_header_row(ws1, len(headers))
    style_data_rows(ws1, len(models), len(headers))
    auto_width(ws1, len(headers))
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(models) + 1}"

    # -------------------------------------------------------------------------
    # Sheet 2: ApexAPI Models
    # -------------------------------------------------------------------------
    ws2 = wb.create_sheet("ApexAPI Models")

    for col, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=h)

    apex_models = [
        # Order
        ("Order", "id", "int", "", "Required", "Apex Trading order ID"),
        ("Order", "uuid", "str", "", "Required", "Apex Trading order UUID"),
        ("Order", "invoice_number", "str", "", "Required", "Invoice number"),
        ("Order", "total", "str", "", "Required", "Order total as string"),
        ("Order", "order_date", "str", "", "Required", "Order date string"),
        ("Order", "order_status", "str", "", "Required", "Status: Pending, Shipped, Delivered, etc."),
        ("Order", "buyer_company", "str", "", "Required", "Buyer company name"),
        ("Order", "seller_company", "str", "", "Required", "Seller company name"),
        ("Order", "ship_name", "str", "", "Required", "Shipping destination name"),
        ("Order", "ship_city", "str", "", "Required", "Shipping city"),
        ("Order", "ship_state", "str", "", "Required", "Shipping state"),
        ("Order", "payment_status", "str", "", "Required", "Payment status"),
        ("Order", "items_count", "int", "", "Required", "Number of line items"),
        ("Order", "printed", "bool", "False", "", "Whether order form has been printed"),
        ("Order", "print_timestamp", "str | None", "None", "", "When the order was printed"),

        # PreRollItem
        ("PreRollItem", "name", "str", "", "Required", "Full product name from Apex"),
        ("PreRollItem", "ordered", "int", "", "Required", "Units ordered"),
        ("PreRollItem", "store_name", "str", "", "Required", "Store that placed the order"),
        ("PreRollItem", "order_id", "int", "", "Required", "Reference to Apex order ID"),
        ("PreRollItem", "order_date", "str", "", "Required", "When the order was placed"),
        ("PreRollItem", "invoice_number", "str", '""', "", "Invoice number for reference"),
        ("PreRollItem", "batch_name", "str", '""', "", "Batch/METRC number from order"),
        ("PreRollItem", "packed", "int", "0", "", "Units packed (ready for delivery)"),
        ("PreRollItem", "synced_packed", "int", "0", "", "Last synced packed count (offline detection)"),
        ("PreRollItem", "synced_ordered", "int", "0", "", "Last synced ordered value"),
        ("PreRollItem", "is_new", "bool", "False", "", "True if order < 24 hours old"),
        ("PreRollItem", "needs_production", "int", "0", "", "Units still needing manufacture"),
        ("PreRollItem", "delivery_date", "str", '""', "", "Delivery date or range"),
        ("PreRollItem", "delivery_notes", "str", '""', "", "Notes visible to stores"),
        ("PreRollItem", "internal_notes", "str", '""', "", "Internal notes (hidden from stores)"),
        ("PreRollItem", "apex_decremented", "bool", "False", "", "True if Apex inventory decremented"),

        # VapeCartProduct
        ("VapeCartProduct", "name", "str", "", "Required", "Full product name"),
        ("VapeCartProduct", "cart_type", "str", "", "Required", "Cart type (0.5g, 1g, 2g, Disposable)"),
        ("VapeCartProduct", "quantity", "int", "", "Required", "Units ordered"),

        # VapeCartSummary
        ("VapeCartSummary", "half_gram", "int", "0", "", "0.5g cartridge count"),
        ("VapeCartSummary", "half_gram_disposable", "int", "0", "", "0.5g disposable count"),
        ("VapeCartSummary", "one_gram", "int", "0", "", "1g cartridge count"),
        ("VapeCartSummary", "one_gram_disposable", "int", "0", "", "1g disposable count"),
        ("VapeCartSummary", "two_gram_disposable", "int", "0", "", "2g disposable count"),
        ("VapeCartSummary", "date_range_start", "str", '""', "", "Start of date range"),
        ("VapeCartSummary", "date_range_end", "str", '""', "", "End of date range"),
        ("VapeCartSummary", "product_breakdown", "list[VapeCartProduct]", "[]", "", "Individual product details"),

        # EdiblesProduct
        ("EdiblesProduct", "name", "str", "", "Required", "Full product name"),
        ("EdiblesProduct", "product_type", "str", "", "Required", "Product type category"),
        ("EdiblesProduct", "category", "str", "", "Required", "Product category"),
        ("EdiblesProduct", "quantity", "int", "", "Required", "Units ordered"),
        ("EdiblesProduct", "pack_size", "str | None", "None", "", "Pack size description"),
        ("EdiblesProduct", "dosage_mg", "int | None", "None", "", "Dosage in milligrams"),
        ("EdiblesProduct", "pack_count", "int | None", "None", "", "Number of items per pack"),
        ("EdiblesProduct", "production_equivalent", "float | None", "None", "", "Production equivalent units"),
        ("EdiblesProduct", "batch_reference", "str | None", "None", "", "Production batch reference"),

        # EdiblesSummary
        ("EdiblesSummary", "date_range_start", "str", '""', "", "Start of analysis period"),
        ("EdiblesSummary", "date_range_end", "str", '""', "", "End of analysis period"),
        ("EdiblesSummary", "product_breakdown", "list[EdiblesProduct]", "[]", "", "Individual product details"),
        ("EdiblesSummary", "production_metrics", "dict | None", "None", "", "Production analysis metrics"),
        ("EdiblesSummary", "kitchen_recommendations", "list[dict] | None", "None", "", "Kitchen workload recommendations"),
        ("EdiblesSummary", "inventory_status", "dict | None", "None", "", "Current inventory status"),

        # PreRollSummary
        ("PreRollSummary", "items", "list[PreRollItem]", "[]", "", "All pre-roll line items"),
        ("PreRollSummary", "filter_type", "str", '"active"', "active/date_range", "Filter mode used"),
        ("PreRollSummary", "date_range_start", "str | None", "None", "", "Start of date range filter"),
        ("PreRollSummary", "date_range_end", "str | None", "None", "", "End of date range filter"),
        ("PreRollSummary", "generated_at", "str", '""', "", "Timestamp when summary was generated"),
    ]

    for row_idx, m in enumerate(apex_models, 2):
        for col_idx, value in enumerate(m, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value)

    style_header_row(ws2, len(headers))
    style_data_rows(ws2, len(apex_models), len(headers))
    auto_width(ws2, len(headers))
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(apex_models) + 1}"

    output_path = "/Users/chrisgillis/PycharmProjects/HiMoM/docs/api-reference/data-models.xlsx"
    wb.save(output_path)
    print(f"Generated: {output_path}")


if __name__ == "__main__":
    generate_api_reference()
    generate_data_models()
